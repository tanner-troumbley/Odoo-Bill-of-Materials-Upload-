from tkinter import *
from tkinter import filedialog
from tkinter import messagebox
from PandasEditor import PandasMagic
from OdooClient import OdooClient
import openpyxl
import logging as log
import traceback
import sys
import warnings

from dotenv import load_dotenv
from os import environ

log.basicConfig(filename='Upload.log', format='%(levelname)s:%(message)s', level=log.DEBUG)
myl = log.getLogger()
warnings.simplefilter("ignore", UserWarning)

load_dotenv()

client = OdooClient(environ['URL'], environ['DATABSE'], environ['USERNAME'], environ['API_KEY'])


class BOM:
    def __init__(self, workbook, client, ITAR=False):
        self.struct = []
        self.client = client
        try:
            self.workbook = openpyxl.load_workbook(workbook)
        except PermissionError:
            log.critical("Close the excel file before uploading it.")
            return
        self.sheet = self.workbook.active
        self.types = self.getTypes()
        self.unreleased = []

        if ITAR or self.sheet.cell(row=1, column=12).value != 'SPEC':
            self.col_map = {
                'level': 2,
                'description': 3,
                'revision': 4,
                'qty': 5,
                'supplier': 6,
                'arrival date': 7,
                'ordered': 8,
                'received': 9,
                'Fab location': 10,
                'Min Stock': 11,
                'mfg pn': 12,
                'Default Approved for': 13,
                'manufacturer': 14,
                'eps part number': 15,
                'comment': 16,
                'type': 17,
                'categ_id': 18,
                'secondary mfg pn': 19,
                'Secondary Approved for': 20,
                'secondary manufacturer': 21,
                "tertiary mfg pn": 22,
                'Tertiary Approved for': 23,
                'tertiary manufacturer': 24,
                'material type': 25,
                'material': 26,
                'finish': 27,
                'weight': 28,
                'volume': 29
            }
        else:
            self.col_map = {
                'level': 2,
                'description': 3,
                'revision': 4,
                'qty': 5,
                'supplier': 6,
                'arrival date': 7,
                'ordered': 8,
                'received': 9,
                'Fab location': 10,
                'Min Stock': 11,
                'SPEC': 12,
                'mfg pn': 13,
                'Default Approved for': 14,
                'manufacturer': 15,
                'eps part number': 16,
                'comment': 17,
                'type': 18,
                'categ_id': 19,
                'secondary mfg pn': 20,
                'Secondary Approved for': 21,
                'secondary manufacturer': 22,
                "tertiary mfg pn": 23,
                'Tertiary Approved for': 24,
                'tertiary manufacturer': 25,
                'material type': 26,
                'material': 27,
                'finish': 28,
                'weight': 29,
                'volume': 30
            }

    def getTypes(self):
        type_dict = {}
        types = self.client.getFields('product.template', ['type'], attrs=['selection'])['type']['selection']
        for typ in types:
            type_dict[typ[0]] = typ[1]
        return type_dict

    def upload(self):
        self.addParts()
        # if len(self.unreleased) > 0:
        #     result = messagebox.askquestion("Choose to Proceed",
        #                                     'This is an unreleased parts. Do you wish to upload anyways?')
        #     log.info(f"Unreleased parts: {self.unreleased}")
        #     if result == 'no':
        #         return
        self.addAll(0, 1, 2)
        self.createBoms()

    def upload_parts(self):
        self.addParts()
        if len(self.unreleased) > 0:
            result = messagebox.askquestion("Choose to Proceed",
                                            'This is an unreleased parts. Do you wish to upload anyways?')
            log.info(f"Unreleased parts: {self.unreleased}")

            if result == 'no':
                return


    def addParts(self):
        partner = self.client.search('res.partner', [('name', '=', 'Electric Power Systems, Inc.')])[0]
        for i in range(2, len(list(self.sheet.rows)) + 1):
            level = str(self.sheet.cell(row=i, column=self.col_map['level']).internal_value)
            level = level.split('.') if level != 'None' and level != '' else []
            self.struct.append([level])

            """Part Number"""
            part = str(self.sheet.cell(row=i, column=self.col_map['eps part number']).internal_value)
            part = False if part == 'None' or part == '' else part.replace(' ', '')
            if not part:
                continue

            """Revision"""
            rev = str(self.sheet.cell(row=i, column=self.col_map['revision']).internal_value)
            rev.replace(' ', '')
            rev.replace('\*', '')
            if "-" in rev or "+" in rev:
                self.unreleased.append(f"{part} REV {rev}")
            rev = False if rev == 'None' or rev == '' else rev
            revision = ' Rev ' + str(rev) if rev else ""

            """Barcode"""
            barcode = str(part) + revision
            barcode = False if not part else barcode
            """QTY"""
            qty = self.sheet.cell(row=i, column=self.col_map['qty']).internal_value
            qty = 1 if qty is None else qty

            """Find existing part or determine on must be created"""
            odooParts = self.client.searchRead('product.template', [('default_code', '=', barcode), ('active', 'in', (True, False))],
                                               ['active', 'default_code', 'version'])

            if len(odooParts) > 0:
                odooPart = odooParts[0]
                if odooPart['active']:
                    """If part exists but is archived, un-archive it"""
                    self.client.update('product.template', odooPart['id'], {'active': True})

            else:
                odooPart = False

            """Create Part if Necessary"""
            if not odooPart:
                description = str(self.sheet.cell(row=i, column=self.col_map['description']).internal_value)
                description = '' if description == 'None' else description.strip()

                """Product Type"""
                prod_type = str(self.sheet.cell(row=i, column=self.col_map['type']).internal_value)
                if prod_type in self.types:
                    prod_type = self.types[prod_type]
                else:
                    prod_type = 'product'

                """Product Category"""
                categ = str(self.sheet.cell(row=i, column=self.col_map['categ_id']).internal_value)
                categ_id = self.client.search('product.category', [['complete_name', '=', categ]])
                categ_id = categ_id[0] if len(categ_id) > 0 else 1

                vals = {'default_code': barcode, 'name': description, 'barcode': barcode, 'version': rev,
                        'tracking': 'lot', 'type': prod_type, 'categ_id': categ_id}

                if 'EWA_H' in barcode or 'Harness' in barcode or 'A0' in barcode or 'ASM' in barcode:
                    vals.update(
                        {'route_ids': self.client.search('stock.route', [('name', '=', 'Manufacture')]), 'purchase_ok': False, 'sale_ok': True})
                else:
                    vals.update({'route_ids': self.client.search('stock.route', [('name', '=', 'Buy')]), 'purchase_ok': True, 'sale_ok': False})

                """Manufacturing Data"""
                # Putting Manufacturers in a list.
                mfgs = [str(self.sheet.cell(row=i, column=self.col_map['manufacturer']).internal_value),
                        str(self.sheet.cell(row=i, column=self.col_map['secondary manufacturer']).internal_value),
                        str(self.sheet.cell(row=i, column=self.col_map['tertiary manufacturer']).internal_value),
                        ]
                # Putting the part numbers in a list with same order.
                mfg_pns = [str(self.sheet.cell(row=i, column=self.col_map['mfg pn']).internal_value),
                           str(self.sheet.cell(row=i, column=self.col_map['secondary mfg pn']).internal_value),
                           str(self.sheet.cell(row=i, column=self.col_map['tertiary mfg pn']).internal_value),
                           ]
                # Loop through and create attributes and the data needed for the product template attributes.
                manufacture_ids = []
                for j in range(len(mfgs)):
                    mfg = mfgs[j] if not mfgs[j] == 'None' else False
                    mfg_pn = mfg_pns[j] if not mfg_pns[j] == 'None' else False
                    if mfg_pn:
                        manufacture_ids.append((0, 0, {'manufacture_name': mfg, 'manufacture_part_number': mfg_pn}))
                try:
                    vals = {'default_code': barcode, 'name': description, 'barcode': barcode,
                            'tracking': 'lot', 'purchase_ok': True, 'sale_ok': False, 'type': prod_type,
                            'categ_id': categ_id, 'manufacture_ids': manufacture_ids, 'seller_ids': [(0, 0, {'delay': 0, 'min_qty': 0, 'partner_id': 5266})],
                            'responsible_id': False}

                    log.debug(f"vals: {vals}")
                    odooPart = {'default_code': barcode, 'version': rev, 'id': self.client.create('product.template', vals)}
                except Exception as e:
                    log.critical(f"Error Uplaoding: {barcode} \nError Details:{e}")

            """Append part to bom structure"""
            self.struct[i - 2].append(odooPart)
            self.struct[i - 2].append(int(qty))
            self.struct[i - 2].append([])

    def add(self, parent, index):
        if len(self.struct[index]) > 1:
            part = self.struct[index][1]
            qty = self.struct[index][2]
            self.struct[parent][3].append([part, qty])

    def addAll(self, parent, index, level):
        if index >= len(self.struct) or index == -1:
            return -1
        depth = len(self.struct[index][0])
        if depth < level:
            return index
        next_row = index + 1
        if depth > level:
            next_row = self.addAll(index - 1, index, depth)
        else:
            self.add(parent, index)
        return self.addAll(parent, next_row, level)

    def createBoms(self):
        for prod in self.struct:
            if len(prod) > 1:
                part = prod[1]
                if len(prod[3]) == 0:
                    continue
                bom = self.client.search('mrp.bom', [['product_tmpl_id', '=', part['id']]])
                if len(bom) > 0:
                    continue
                bom_lines = []
                lines = prod[3]
                for line in lines:
                    component = line[0]
                    try:
                        odooPart = self.client.search('product.product', [('product_tmpl_id', '=', component['id'])])[0]
                        if not odooPart:
                            raise Exception
                        qty = line[1]
                        bom_lines.append((0, 0, {'product_id': odooPart, 'product_qty': qty}))
                    except TypeError:
                        log.critical(f"{component} is missing the product.product record id.")
                        continue
                log.debug(f" part: {part}")
                log.debug(f" bom_lines: {bom_lines}")
                self.client.create('mrp.bom', {'product_tmpl_id': part['id'], 'code': f"Rev {part['version']}", 'bom_line_ids': bom_lines})
                self.client.update('product.template', part['id'], {"route_ids": [5]})


class Upload:
    def choose_file(self):
        gui = Tk()
        gui.withdraw()
        file_path = filedialog.askopenfilename(multiple=True)
        return file_path

    def main(self):
        files = self.choose_file()
        count = 1
        for data in files:
            try:
                log.info(f"{count}/{len(files)}")
                bom = BOM(data, client)
                bom.upload()
                log.info(f"Uploaded {data.split('/')[-1]}")
                count += 1
            except Exception as e:
                print("Error Check log")
                log.critical(f"Uncaught exception in main loop {e}")
                log.critical(traceback.format_exc())
                log.critical(sys.exc_info()[2])
        print("Uploads Finished")

    def main_itar(self):
        files = self.choose_file()
        count = 1
        for data in files:
            try:
                log.info(f"{count}/{len(files)}")
                bom = BOM(data, client, ITAR=True)
                bom.upload()
                log.info(f"Uploaded {data.split('/')[-1]}")
                count += 1
            except Exception as e:
                print("Error Check log")
                log.critical(f"Uncaught exception in main loop {e}")
                log.critical(traceback.format_exc())
                log.critical(sys.exc_info()[2])
        print("Uploads Finished")

    def upload_parts(self):
        files = self.choose_file()
        for data in files:
            try:
                bom = BOM(data, client)
                bom.upload_parts()
                log.info(f"Uploaded {data.split('/')[-1]}")
            except Exception as e:
                print("Error Check log")
                log.critical(f"Uncaught exception in main loop {e}")
                log.critical(traceback.format_exc())
                log.critical(sys.exc_info()[2])
        print("Uploads Finished")


class TextRedirector(object):
    def __init__(self, widget, tag="stdout"):
        self.widget = widget
        self.tag = tag

    def write(self, str):
        self.widget.configure(state="normal")
        self.widget.insert("end", str, (self.tag,))
        self.widget.see('end')
        self.widget.configure(state="disabled")


upload = Upload()
edit = PandasMagic()
# Open's window
home = Tk()
home.geometry('800x600')
home['bg'] = '#887c87'
home.title('Bomb Editor/Uploader')

# creates buttons
btnUpload = Button(home, text="Upload Part File.", bd=5, command=upload.upload_parts)
btnUpload.place(relx=0.36, y=0)

btnUpload = Button(home, text="Upload ITAR BOM.", bd=5, command=upload.main_itar)
btnUpload.place(relx=0.2, y=0)

btnUpload = Button(home, text="Upload Bom File.", bd=5, command=upload.main)
btnUpload.place(relx=0.5, y=0)

btnExit = Button(home, text='Exit Program', bd=1, command=exit)
btnExit.place(relx=1.0, y=0, anchor='ne')

btnEdit = Button(home, text="Edit File", command=edit.main)
btnEdit.place(relx=0, y=0)
home.mainloop()
